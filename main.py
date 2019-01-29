year_start = 2017
year_end = 2018
month_start = 4
month_end = 11
file_name = 'data.xlsx'

import requests
import utils
import numpy

from bs4 import BeautifulSoup
from openpyxl import Workbook

wb = Workbook()
sheet = wb.active

x_data, y_data = [],[]

for year in range(year_start, year_end + 1):
    for month in range(month_start, month_end + 1):
        for day in range(1 , 32):
            print('Year %d Month %d  Day %d' %(year, month, day))
            url = 'http://www.scorecenterlive.com/ko/sports-livescore.html? sports=bs&nation=ko&date='+str(year)+'-'+utils.numtoten(month)+'-'+utils.numtoten(day)
    
            source_code = requests.get(url).text
            soup = BeautifulSoup(source_code, 'lxml')
    
            name, score = [], []
            min, max = -1, -1
            index = 0
    
            #Parsing Data
            for data in soup.find_all("li", {"class":"team_n"}):
                if int(utils.KBO_num(data.string)) < 10:
                    if min == -1:
                        min = index
                    max = index
                    name.append(data.string)
                index = index + 1
    
            index = 0
            index2 = 0 #name의 value를 지우기 위한 index
            for data in soup.find_all("li", {"class":"score"}):
                if index >= min and index <= max and len(name) > 1:
                    if data.string == None:
                        name[index2] = ""
                    else:
                        score.append(int(data.string))
                    index2 = index2 + 1
                index = index + 1
    
            #Processing Data
            index = len(name) - 1
            while index >= 0:
                if name[index] == "":
                    del name[index]
                index = index - 1
    
            if len(score) > 1:
                for i in range(len(name)):
                    name[i] = utils.KBO_num(name[i])
    
                tmp = []
                for i in range(0, len(score), 2):
                    if score[i] > score[i + 1]:
                        tmp.append(0)
                    elif score[i] < score[i + 1]:
                        tmp.append(1)
                    else:
                        tmp.append(2)
    
                score = tmp
                print(name, score)
    
                x_data.extend(name)
                y_data.extend(score)
    
    x_data = numpy.reshape(numpy.array(x_data), [int(len(x_data)/2), 2])
    
    for i in range(x_data.shape[0]):
        sheet.cell(row=i+1, column=1, value=x_data[i, 0])
        sheet.cell(row=i+1, column=2, value=x_data[i, 1])
        sheet.cell(row=i+1, column=3, value=y_data[i])

wb.save(filename=file_name)
